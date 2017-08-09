import openpyxl
import re
from Customer import CustomerDevice

#Look at column G, if the model contains '3560' or '2960',
#If it does, create a CustomerDevice, which is child of the device_class

#Proprety Name (B)
#WiFi Type (C)
#Granite CLLI (K)
#unique Property ID (L)

#This class might be called something like customer, and contain a list of devices
#Those devices wil have specifics like:

#Device Name (D
#Device Function (E)
#IP (F)
#Design ID (H)
#username
#password



#for each row in column G,check to see if it's says
#'3560' or '2960'
#if it does, then create a new customer object and populate
#the necessary fields, then create a CustomerDevice object
#and populate those sepcific fields.

##this might need to be a "has-a" type relationship and not an "is-a" child
#relationship

def main(file):
    device_list = []

    wb = openpyxl.load_workbook(file)
    sheet = wb.active

    for col in sheet.iter_rows(min_row=1,max_col=1,max_row=sheet.max_row,column_offset=6):
        for cell in col:
            if cell.value == None:
                continue
            if re.search('3560', cell.value) != None or re.search('2960', cell.value) != None:
                row = str(cell.row)
                name = sheet['D'+row].value
                function = sheet['E'+row].value
                ip = sheet['F'+row].value
                design_id = sheet['H'+row].value
                #print("Row: ", row, cell.value, " is a switch")
                device = CustomerDevice(ip,'','',100000,row, name, function, design_id)
                device_list.append(device)
                

    for device in device_list:
        print("-----------------------")
        print('Row: ', device.row)
        print(device.name)
        print(device.ip)
        print(device.function)
        print(device.design_id)
        print("-----------------------")


if __name__ == '__main__':
    main('Test.xlsx')
            



